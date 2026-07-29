"""Импорт каталога из Excel-шаблона сметы Intro Show («Для нас»)."""
from __future__ import annotations

from typing import Any, Dict, List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from sqlalchemy.orm import Session


def _is_section_header(col_b: Any, col_c: Any) -> bool:
    b = str(col_b).lower() if col_b is not None else ""
    c = str(col_c).lower() if col_c is not None else ""
    return "кол-во" in b or "цена" in c


def _is_meta_row(name: str) -> bool:
    low = name.lower()
    keys = (
        "наименование проекта",
        "контактн",
        "менеджер",
        "город",
        "дата проведения",
        "количество дней",
        "выезд оборудования",
        "возврат оборудования",
        "фактический возврат",
        "итог",
        "скидк",
        "налог",
        "марж",
        "себестоим",
        "субаренд",
        "только для нас",
    )
    return any(k in low for k in keys)


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(" ", "").replace(",", ".")
        return float(s)
    except (TypeError, ValueError):
        return None


def _to_int(v: Any) -> int:
    f = _to_float(v)
    if f is None:
        return 0
    return max(0, int(round(f)))


def parse_estimate_xlsx(file_path: str, sheet_name: str = "Для нас") -> List[Dict[str, Any]]:
    """Парсит лист «Для нас»: секции + позиции (name, price, stock, power_kw)."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows: List[Dict[str, Any]] = []
    section = "Без категории"

    for r in range(1, (ws.max_row or 0) + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        f = ws.cell(r, 6).value  # Кол-во на базе
        g = ws.cell(r, 7).value  # кВт
        if a is None:
            continue
        name = str(a).strip().replace("\xa0", " ")
        if not name:
            continue
        if _is_section_header(b, c):
            section = name
            continue
        if _is_meta_row(name):
            continue
        price = _to_float(c)
        if price is None:
            continue
        rows.append({
            "section": section,
            "name": name,
            "price": price,
            "stock_quantity": _to_int(f),
            "power_kw": _to_float(g),
        })
    return rows


def import_catalog_from_xlsx(
    db: "Session",
    file_path: str,
    *,
    update_existing: bool = True,
    sheet_name: str = "Для нас",
) -> Dict[str, Any]:
    """Создаёт/обновляет Folder + Equipment из Excel. Не удаляет чужие позиции."""
    from database import Equipment, Folder

    parsed = parse_estimate_xlsx(file_path, sheet_name=sheet_name)
    created = 0
    updated = 0
    folders_created = 0
    folder_cache: Dict[str, Folder] = {}

    for row in parsed:
        sec = (row["section"] or "Без категории").strip()
        if sec not in folder_cache:
            folder = db.query(Folder).filter(Folder.name == sec).first()
            if not folder:
                folder = Folder(name=sec)
                db.add(folder)
                db.flush()
                folders_created += 1
            folder_cache[sec] = folder
        folder = folder_cache[sec]

        eq = db.query(Equipment).filter(Equipment.name == row["name"]).first()
        power_w = None
        if row.get("power_kw") is not None:
            power_w = float(row["power_kw"]) * 1000.0  # кВт → Вт

        if eq:
            if not update_existing:
                continue
            eq.price = row["price"]
            eq.category = sec
            eq.folder_id = folder.id
            if row["stock_quantity"]:
                eq.stock_quantity = row["stock_quantity"]
            if power_w is not None:
                eq.power_w = power_w
            if not getattr(eq, "warehouse_type", None):
                eq.warehouse_type = "own"
            updated += 1
        else:
            eq = Equipment(
                name=row["name"],
                category=sec,
                price=row["price"],
                stock_quantity=row["stock_quantity"] or 0,
                folder_id=folder.id,
                power_w=power_w,
                warehouse_type="own",
                status="Доступно",
            )
            db.add(eq)
            created += 1

    db.commit()
    return {
        "status": "success",
        "parsed": len(parsed),
        "created": created,
        "updated": updated,
        "folders_created": folders_created,
        "sections": len(folder_cache),
    }
